import openpyxl
import pyperclip
import pyautogui
from time import sleep
# Entrar na planilha

workbook = openpyxl.load_workbook('produtos_ficticios.xlsx')
# Selecionar apenas a pagina
sheet_produtos = workbook['Produtos']
# Copiar informação de um campo e colar no campo respectivo

for linha in sheet_produtos.iter_rows(min_row=2): # iter_rows = passa sobre cada linha de acordo com uma especificação
    nome_produto = linha[0].value
    pyperclip.copy(nome_produto)
    pyautogui.click(1528,336,duration=1)
    pyautogui.hotkey('ctrl','v')

    descricao = linha[1].value
    pyperclip.copy(descricao)
    pyautogui.hotkey('tab')
    pyautogui.hotkey('ctrl','v')

    categoria = linha[2].value
    pyperclip.copy(categoria)
    pyautogui.hotkey('tab')
    pyautogui.hotkey('ctrl','v')

    codigo = linha[3].value
    pyperclip.copy(codigo)
    pyautogui.hotkey('tab')
    pyautogui.hotkey('ctrl','v')

    peso = linha[4].value
    pyperclip.copy(peso)
    pyautogui.hotkey('tab')
    pyautogui.hotkey('ctrl','v')

    dimensoes = linha[5].value
    pyperclip.copy(dimensoes)
    pyautogui.hotkey('tab')
    pyautogui.hotkey('ctrl','v')
    # Apertar em proximo
    pyautogui.hotkey('tab')
    pyautogui.hotkey('enter')
    sleep(3) # Pausa para garantir que a proxima pagina vai estar carregada

    preco = linha[6].value
    pyperclip.copy(preco)
    pyautogui.hotkey('tab')
    pyautogui.hotkey('ctrl','v')

    quant_estoque = linha[7].value
    pyperclip.copy(quant_estoque)
    pyautogui.hotkey('tab')
    pyautogui.hotkey('ctrl','v')

    data_validade = linha[8].value
    pyperclip.copy(data_validade)
    pyautogui.hotkey('tab')
    pyautogui.hotkey('ctrl','v')

    cor = linha[9].value
    pyperclip.copy(cor)
    pyautogui.hotkey('tab')
    pyautogui.hotkey('ctrl','v')

    tamanho = linha[10].value
    pyautogui.hotkey('tab')
    pyautogui.hotkey('enter')
    if tamanho == 'Pequeno':
        pyautogui.click(1288,756,duration=1)
    elif tamanho == 'Médio':
        pyautogui.click(1188,773,duration=1)
    else:
        pyautogui.click(1221,797,duration=1)
    
    material = linha[11].value
    pyperclip.copy(cor)
    pyautogui.hotkey('tab')
    pyautogui.hotkey('ctrl','v')

    pyautogui.hotkey('tab')
    pyautogui.hotkey('enter')
    sleep(3)

    fabricante = linha[12].value
    pyperclip.copy(fabricante)
    pyautogui.hotkey('tab')
    pyautogui.hotkey('ctrl','v')

    pais_origem = linha[13].value
    pyperclip.copy(pais_origem)
    pyautogui.hotkey('tab')
    pyautogui.hotkey('ctrl','v')

    observacoes = linha[14].value
    pyperclip.copy(observacoes)
    pyautogui.hotkey('tab')
    pyautogui.hotkey('ctrl','v')

    cod_barra = linha[15].value
    pyperclip.copy(cod_barra)
    pyautogui.hotkey('tab')
    pyautogui.hotkey('ctrl','v')

    localizacao_armazem = linha[16].value
    pyperclip.copy(localizacao_armazem)
    pyautogui.hotkey('tab')
    pyautogui.hotkey('ctrl','v')
    
    # Botão concluir
    pyautogui.hotkey('tab')
    pyautogui.hotkey('enter')
    sleep(2)
    # Botão confirmar inclusão
    pyautogui.hotkey('enter')
    sleep(2)
    # Botão ok
    pyautogui.hotkey('enter')
    sleep(2)
    # Iniciar mais um cadastro
    pyautogui.hotkey('tab')
    pyautogui.hotkey('enter')
    sleep(2)